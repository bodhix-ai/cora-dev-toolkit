"""
Eval Results Lambda - Evaluation CRUD, Edits, Export

Handles evaluation creation, retrieval, result editing, and report export.
Triggers async processing via SQS for new evaluations.

Routes - Evaluation CRUD:
- POST /workspaces/{wsId}/eval - Create evaluation
- GET /workspaces/{wsId}/eval - List evaluations
- GET /workspaces/{wsId}/eval/{id} - Get evaluation detail
- PATCH /workspaces/{wsId}/eval/{id} - Update draft evaluation (configure and trigger processing)
- GET /workspaces/{wsId}/eval/{id}/status - Get progress status
- DELETE /workspaces/{wsId}/eval/{id} - Delete evaluation

Routes - Result Editing:
- PATCH /workspaces/{wsId}/eval/{id}/results/{resultId} - Edit result
- GET /workspaces/{wsId}/eval/{id}/results/{resultId}/history - Get edit history

Routes - Export:
- GET /workspaces/{wsId}/eval/{id}/export/pdf - Export PDF
- GET /workspaces/{wsId}/eval/{id}/export/xlsx - Export XLSX
"""

import json
import logging
import os
import io
import base64
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional
from decimal import Decimal

# Add the layer to path
import sys
sys.path.insert(0, '/opt/python')

import org_common as common
import boto3

# Configure logging
LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO")
logger = logging.getLogger()
logger.setLevel(LOG_LEVEL)

# SQS Configuration
SQS_QUEUE_URL = os.getenv("EVAL_PROCESSOR_QUEUE_URL", "")
AWS_REGION = os.getenv("AWS_REGION", "us-east-1")

# Export Configuration
EXPORT_BUCKET = os.getenv("EXPORT_BUCKET", "")


def lambda_handler(event: Dict[str, Any], context: Any) -> Dict[str, Any]:
    """Main Lambda handler for evaluation operations."""
    
    logger.info(f"Eval Results event: {json.dumps(event, default=str)}")
    
    try:
        # Extract user info from authorizer
        user_info = common.get_user_from_event(event)
        okta_uid = user_info['user_id']
        
        # Convert Okta UID to Supabase UUID
        supabase_user_id = common.get_supabase_user_id_from_okta_uid(okta_uid)
        
        # Extract HTTP method and path
        http_method = event.get('requestContext', {}).get('http', {}).get('method') or event.get('httpMethod')
        if not http_method:
            return common.bad_request_response('HTTP method not found in request')
        
        path = event.get('rawPath', '') or event.get('path', '')
        path_params = event.get('pathParameters', {}) or {}
        
        # Handle OPTIONS for CORS
        if http_method == 'OPTIONS':
            return common.success_response({})
        
        # Extract workspace_id from path
        ws_id = path_params.get('wsId')
        if not ws_id:
            return common.bad_request_response('Workspace ID is required')
        
        ws_id = common.validate_uuid(ws_id, 'wsId')
        
        # Verify workspace access
        if not has_workspace_access(supabase_user_id, ws_id):
            return common.forbidden_response('Workspace access denied')
        
        # Get org_id for workspace
        workspace = common.find_one('workspaces', {'id': ws_id})
        if not workspace:
            return common.not_found_response('Workspace not found')
        
        org_id = workspace.get('org_id')
        
        # =================================================================
        # EVALUATION CRUD ROUTES
        # =================================================================
        
        eval_id = path_params.get('evalId')
        result_id = path_params.get('resultId')
        
        # Export routes
        if '/export/pdf' in path and http_method == 'GET':
            return handle_export_pdf(eval_id, ws_id, org_id, supabase_user_id)
        
        if '/export/xlsx' in path and http_method == 'GET':
            return handle_export_xlsx(eval_id, ws_id, org_id, supabase_user_id)
        
        # Result editing routes
        if result_id:
            if '/history' in path and http_method == 'GET':
                return handle_get_edit_history(eval_id, result_id, ws_id, supabase_user_id)
            elif http_method == 'PATCH':
                return handle_edit_result(event, eval_id, result_id, ws_id, supabase_user_id)
        
        # Status route
        if eval_id and '/status' in path and http_method == 'GET':
            return handle_get_status(eval_id, ws_id)
        
        # Main evaluation CRUD
        if eval_id:
            if http_method == 'GET':
                return handle_get_evaluation(eval_id, ws_id, org_id)
            elif http_method == 'PATCH':
                return handle_update_evaluation(event, eval_id, ws_id, org_id, supabase_user_id)
            elif http_method == 'DELETE':
                return handle_delete_evaluation(eval_id, ws_id, supabase_user_id)
        else:
            if http_method == 'POST':
                return handle_create_evaluation(event, ws_id, org_id, supabase_user_id)
            elif http_method == 'GET':
                return handle_list_evaluations(event, ws_id)
        
        return common.not_found_response('Route not found')
        
    except KeyError as e:
        logger.error(f'KeyError: {str(e)}')
        return common.unauthorized_response(f'Missing user information: {str(e)}')
    except common.NotFoundError as e:
        return common.not_found_response(str(e))
    except common.ValidationError as e:
        return common.bad_request_response(str(e))
    except common.ForbiddenError as e:
        return common.forbidden_response(str(e))
    except Exception as e:
        logger.exception(f'Error: {str(e)}')
        return common.internal_error_response('Internal server error')


# =============================================================================
# AUTHORIZATION HELPERS
# =============================================================================

def has_workspace_access(user_id: str, workspace_id: str) -> bool:
    """Check if user has access to workspace."""
    try:
        # Check workspace membership
        membership = common.find_one(
            'ws_members',
            {'user_id': user_id, 'ws_id': workspace_id}
        )
        if membership:
            return True
        
        # Check if sys admin
        profile = common.find_one('user_profiles', {'user_id': user_id})
        if profile and profile.get('sys_role') in ['sys_owner', 'sys_admin']:
            return True
        
        # Check if org admin for workspace's org
        workspace = common.find_one('workspaces', {'id': workspace_id})
        if workspace:
            org_membership = common.find_one(
                'org_members',
                {'user_id': user_id, 'org_id': workspace['org_id']}
            )
            if org_membership and org_membership.get('org_role') in ['org_owner', 'org_admin']:
                return True
        
        return False
    except Exception as e:
        logger.error(f"Error checking workspace access: {e}")
        return False


def is_eval_owner(user_id: str, eval_id: str) -> bool:
    """Check if user owns the evaluation."""
    try:
        evaluation = common.find_one('eval_doc_summaries', {'id': eval_id})
        return evaluation and evaluation.get('created_by') == user_id
    except Exception as e:
        logger.error(f"Error checking eval ownership: {e}")
        return False


# =============================================================================
# EVALUATION CRUD HANDLERS
# =============================================================================

def handle_create_evaluation(
    event: Dict[str, Any],
    workspace_id: str,
    org_id: str,
    user_id: str
) -> Dict[str, Any]:
    """
    Create a new evaluation and trigger async processing.
    
    Request body (DRAFT MODE - only name required):
    {
        "name": "Evaluation Name"
    }
    
    OR (FULL MODE - create and process immediately):
    {
        "name": "Evaluation Name",
        "docTypeId": "uuid",
        "criteriaSetId": "uuid",
        "docIds": ["uuid1", "uuid2"]
    }
    """
    body = json.loads(event.get('body', '{}'))
    
    # Validate required fields
    name = body.get('name')
    if not name:
        raise common.ValidationError('name is required')
    name = common.validate_string_length(name, 'name', max_length=200)
    
    # Check if this is draft mode (missing config fields)
    doc_type_id = body.get('docTypeId')
    criteria_set_id = body.get('criteriaSetId')
    doc_ids = body.get('docIds', [])
    
    is_draft = not doc_type_id or not criteria_set_id or not doc_ids
    
    if is_draft:
        # DRAFT MODE: Create incomplete evaluation for later configuration
        evaluation_data = {
            'ws_id': workspace_id,
            'doc_type_id': None,
            'criteria_set_id': None,
            'name': name,
            'status': 'draft',
            'progress': 0,
            'is_deleted': False,
            'created_by': user_id,
            'updated_by': user_id
        }
        
        evaluation = common.insert_one('eval_doc_summaries', evaluation_data)
        
        # Return draft evaluation (no doc type or criteria set info)
        result = common.format_record(evaluation)
        result['docType'] = None
        result['criteriaSet'] = None
        result['docIds'] = []
        
        return common.created_response(result)
    
    # FULL MODE: Validate all required fields
    doc_type_id = common.validate_uuid(doc_type_id, 'docTypeId')
    criteria_set_id = common.validate_uuid(criteria_set_id, 'criteriaSetId')
    
    # Validate doc_ids are UUIDs
    validated_doc_ids = []
    for doc_id in doc_ids:
        validated_doc_ids.append(common.validate_uuid(doc_id, 'docId'))
    
    # Verify doc type exists and belongs to org
    doc_type = common.find_one('eval_doc_types', {'id': doc_type_id, 'org_id': org_id})
    if not doc_type:
        raise common.NotFoundError('Document type not found')
    
    if not doc_type.get('is_active'):
        raise common.ValidationError('Document type is inactive')
    
    # Verify criteria set exists and belongs to doc type
    criteria_set = common.find_one('eval_criteria_sets', {'id': criteria_set_id, 'doc_type_id': doc_type_id})
    if not criteria_set:
        raise common.NotFoundError('Criteria set not found')
    
    if not criteria_set.get('is_active'):
        raise common.ValidationError('Criteria set is inactive')
    
    # Get workspace KB for document verification
    workspace_kb = common.find_one('kb_bases', {'ws_id': workspace_id})
    if not workspace_kb:
        raise common.NotFoundError('Workspace knowledge base not found')
    
    # Verify documents exist in workspace KB
    for doc_id in validated_doc_ids:
        doc = common.find_one('kb_docs', {'id': doc_id, 'kb_id': workspace_kb['id']})
        if not doc:
            raise common.NotFoundError(f'Document not found: {doc_id}')
    
    # Create evaluation record
    evaluation_data = {
        'ws_id': workspace_id,
        'doc_type_id': doc_type_id,
        'criteria_set_id': criteria_set_id,
        'name': name,
        'status': 'pending',
        'progress': 0,
        'is_deleted': False,
        'created_by': user_id,
        'updated_by': user_id
    }
    
    evaluation = common.insert_one('eval_doc_summaries', evaluation_data)
    eval_id = evaluation['id']
    
    # Create doc set entries
    for idx, doc_id in enumerate(validated_doc_ids):
        common.insert_one('eval_doc_sets', {
            'eval_summary_id': eval_id,
            'kb_doc_id': doc_id,
            'order_index': idx,
            'is_primary': idx == 0
        })
    
    # Send SQS message to trigger processing
    send_processing_message(
        eval_id=eval_id,
        org_id=org_id,
        workspace_id=workspace_id,
        doc_ids=validated_doc_ids,
        criteria_set_id=criteria_set_id
    )
    
    # Return created evaluation
    result = common.format_record(evaluation)
    result['docType'] = common.format_record(doc_type)
    result['criteriaSet'] = {
        'id': criteria_set['id'],
        'name': criteria_set['name'],
        'version': criteria_set.get('version')
    }
    result['docIds'] = validated_doc_ids
    
    return common.created_response(result)


def handle_list_evaluations(event: Dict[str, Any], workspace_id: str) -> Dict[str, Any]:
    """List evaluations for a workspace with pagination and filters."""
    query_params = event.get('queryStringParameters', {}) or {}
    
    # Pagination
    limit = common.validate_integer(
        query_params.get('limit', 20),
        'limit', min_value=1, max_value=100
    )
    offset = common.validate_integer(
        query_params.get('offset', 0),
        'offset', min_value=0
    )
    
    # Filters
    status = query_params.get('status')
    doc_type_id = query_params.get('docTypeId')
    include_deleted = query_params.get('includeDeleted', 'false').lower() == 'true'
    
    # Build filters
    filters = {'ws_id': workspace_id}
    
    if not include_deleted:
        filters['is_deleted'] = False
    
    if status and status in ['pending', 'processing', 'completed', 'failed']:
        filters['status'] = status
    
    if doc_type_id:
        filters['doc_type_id'] = common.validate_uuid(doc_type_id, 'docTypeId')
    
    # Query evaluations
    evaluations = common.find_many(
        'eval_doc_summaries',
        filters,
        order='created_at.desc',
        limit=limit,
        offset=offset
    )
    
    # Get total count for pagination
    total_count = len(common.find_many('eval_doc_summaries', filters, select='id'))
    
    # Format results with doc type info
    results = []
    for eval_record in evaluations:
        formatted = common.format_record(eval_record)
        
        # Get doc type name
        doc_type = common.find_one('eval_doc_types', {'id': eval_record['doc_type_id']})
        if doc_type:
            formatted['docTypeName'] = doc_type['name']
        
        # Get criteria set name
        criteria_set = common.find_one('eval_criteria_sets', {'id': eval_record['criteria_set_id']})
        if criteria_set:
            formatted['criteriaSetName'] = criteria_set['name']
        
        # Get document count
        doc_count = len(common.find_many(
            'eval_doc_sets',
            {'eval_summary_id': eval_record['id']},
            select='id'
        ))
        formatted['documentCount'] = doc_count
        
        results.append(formatted)
    
    return common.success_response({
        'evaluations': results,
        'pagination': {
            'limit': limit,
            'offset': offset,
            'total': total_count,
            'hasMore': offset + len(results) < total_count
        }
    })


def handle_get_evaluation(eval_id: str, workspace_id: str, org_id: str) -> Dict[str, Any]:
    """Get detailed evaluation with all results."""
    eval_id = common.validate_uuid(eval_id, 'id')
    
    # Get evaluation
    evaluation = common.find_one('eval_doc_summaries', {'id': eval_id, 'ws_id': workspace_id})
    if not evaluation:
        raise common.NotFoundError('Evaluation not found')
    
    result = common.format_record(evaluation)
    
    # Get doc type
    doc_type = common.find_one('eval_doc_types', {'id': evaluation['doc_type_id']})
    if doc_type:
        result['docType'] = common.format_record(doc_type)
    
    # Get criteria set
    criteria_set = common.find_one('eval_criteria_sets', {'id': evaluation['criteria_set_id']})
    if criteria_set:
        result['criteriaSet'] = common.format_record(criteria_set)
    
    # Get documents
    doc_sets = common.find_many(
        'eval_doc_sets',
        {'eval_summary_id': eval_id},
        order='order_index.asc'
    )
    
    documents = []
    for doc_set in doc_sets:
        doc = common.find_one('kb_docs', {'id': doc_set['kb_doc_id']})
        if doc:
            documents.append({
                'id': doc['id'],
                'name': doc.get('name'),
                'fileName': doc.get('file_name'),
                'mimeType': doc.get('mime_type'),
                'summary': doc_set.get('doc_summary'),
                'isPrimary': doc_set.get('is_primary', False)
            })
    result['documents'] = documents
    
    # Get criteria results with current edits
    criteria_items = common.find_many(
        'eval_criteria_items',
        {'criteria_set_id': evaluation['criteria_set_id']},
        order='order_index.asc'
    )
    
    criteria_results = common.find_many(
        'eval_criteria_results',
        {'eval_summary_id': eval_id}
    )
    
    results_map = {r['criteria_item_id']: r for r in criteria_results}
    
    # Get effective eval config for this org
    eval_config = get_effective_eval_config(org_id)
    categorical_mode = eval_config.get('categorical_mode', 'detailed')
    show_numerical_score = eval_config.get('show_numerical_score', True)
    
    # Get status options filtered by categorical mode
    status_options = get_status_options(org_id, categorical_mode)
    status_map = {s['id']: s for s in status_options}
    
    criteria_result_list = []
    for item in criteria_items:
        ai_result = results_map.get(item['id'], {})
        
        # Get current edit if exists
        current_edit = None
        if ai_result.get('id'):
            edits = common.find_many(
                'eval_result_edits',
                {'criteria_result_id': ai_result['id'], 'is_current': True},
                limit=1
            )
            if edits:
                current_edit = common.format_record(edits[0])
        
        # Determine effective status
        effective_status_id = (
            current_edit.get('editedStatusId') if current_edit
            else ai_result.get('ai_status_id')
        )
        effective_status = status_map.get(effective_status_id, {})
        
        criteria_result_list.append({
            'criteriaItem': {
                'id': item['id'],
                'criteriaId': item.get('criteria_id'),
                'requirement': item.get('requirement'),
                'description': item.get('description'),
                'category': item.get('category'),
                'weight': float(item.get('weight', 1.0))
            },
            'aiResult': {
                'id': ai_result.get('id'),
                'result': ai_result.get('ai_result'),
                'statusId': ai_result.get('ai_status_id'),
                'scoreValue': ai_result.get('ai_score_value'),
                'confidence': ai_result.get('ai_confidence'),
                'citations': json.loads(ai_result.get('ai_citations', '[]')) if ai_result.get('ai_citations') else [],
                'processedAt': ai_result.get('processed_at')
            } if ai_result else None,
            'currentEdit': current_edit,
            'effectiveStatus': {
                'id': effective_status.get('id'),
                'name': effective_status.get('name'),
                'color': effective_status.get('color'),
                'scoreValue': effective_status.get('score_value')
            } if effective_status else None,
            'hasEdit': current_edit is not None
        })
    
    result['criteriaResults'] = criteria_result_list
    
    # Add status options for frontend (DEPRECATED - use scoreConfig instead)
    result['statusOptions'] = [
        {
            'id': s['id'],
            'name': s['name'],
            'color': s.get('color'),
            'scoreValue': s.get('score_value')
        }
        for s in status_options
    ]
    
    # Add scoreConfig object for configuration-based score display
    # Map hex colors to MUI color names
    color_map = {
        '#4caf50': 'success',
        '#2196f3': 'info',
        '#ff9800': 'warning',
        '#ffeb3b': 'warning',
        '#f44336': 'error',
        '#8bc34a': 'success'
    }
    
    result['scoreConfig'] = {
        'categoricalMode': categorical_mode,
        'showDecimalScore': show_numerical_score,  # Frontend expects 'showDecimalScore'
        'statusOptions': [
            {
                'id': s['id'],
                'name': s['name'],
                'color': color_map.get(s.get('color', ''), 'default'),  # Convert hex to MUI color
                'scoreValue': float(s['score_value']) if s.get('score_value') is not None else None
            }
            for s in status_options
        ]
    }
    
    return common.success_response(result)


def handle_get_status(eval_id: str, workspace_id: str) -> Dict[str, Any]:
    """Get evaluation status and progress (for polling)."""
    eval_id = common.validate_uuid(eval_id, 'id')
    
    evaluation = common.find_one('eval_doc_summaries', {'id': eval_id, 'ws_id': workspace_id})
    if not evaluation:
        raise common.NotFoundError('Evaluation not found')
    
    return common.success_response({
        'id': eval_id,
        'status': evaluation.get('status'),
        'progress': evaluation.get('progress', 0),
        'errorMessage': evaluation.get('error_message'),
        'startedAt': evaluation.get('started_at'),
        'completedAt': evaluation.get('completed_at')
    })


def handle_update_evaluation(
    event: Dict[str, Any],
    eval_id: str,
    workspace_id: str,
    org_id: str,
    user_id: str
) -> Dict[str, Any]:
    """
    Update a draft evaluation with configuration and trigger processing.
    
    Request body:
    {
        "docTypeId": "uuid",
        "criteriaSetId": "uuid",
        "docIds": ["uuid1", "uuid2"]
    }
    """
    eval_id = common.validate_uuid(eval_id, 'id')
    
    # Verify evaluation exists and belongs to workspace
    evaluation = common.find_one('eval_doc_summaries', {'id': eval_id, 'ws_id': workspace_id})
    if not evaluation:
        raise common.NotFoundError('Evaluation not found')
    
    # Verify evaluation is in draft status
    if evaluation.get('status') != 'draft':
        raise common.ValidationError('Only draft evaluations can be updated')
    
    body = json.loads(event.get('body', '{}'))
    
    # Validate required fields
    doc_type_id = body.get('docTypeId')
    criteria_set_id = body.get('criteriaSetId')
    doc_ids = body.get('docIds', [])
    
    if not doc_type_id or not criteria_set_id or not doc_ids:
        raise common.ValidationError('docTypeId, criteriaSetId, and docIds are required')
    
    doc_type_id = common.validate_uuid(doc_type_id, 'docTypeId')
    criteria_set_id = common.validate_uuid(criteria_set_id, 'criteriaSetId')
    
    # Validate doc_ids are UUIDs
    validated_doc_ids = []
    for doc_id in doc_ids:
        validated_doc_ids.append(common.validate_uuid(doc_id, 'docId'))
    
    # Verify doc type exists and belongs to org
    doc_type = common.find_one('eval_doc_types', {'id': doc_type_id, 'org_id': org_id})
    if not doc_type:
        raise common.NotFoundError('Document type not found')
    
    if not doc_type.get('is_active'):
        raise common.ValidationError('Document type is inactive')
    
    # Verify criteria set exists and belongs to doc type
    criteria_set = common.find_one('eval_criteria_sets', {'id': criteria_set_id, 'doc_type_id': doc_type_id})
    if not criteria_set:
        raise common.NotFoundError('Criteria set not found')
    
    if not criteria_set.get('is_active'):
        raise common.ValidationError('Criteria set is inactive')
    
    # Get workspace KB for document verification
    workspace_kb = common.find_one('kb_bases', {'ws_id': workspace_id})
    if not workspace_kb:
        raise common.NotFoundError('Workspace knowledge base not found')
    
    # Verify documents exist in workspace KB
    for doc_id in validated_doc_ids:
        doc = common.find_one('kb_docs', {'id': doc_id, 'kb_id': workspace_kb['id']})
        if not doc:
            raise common.NotFoundError(f'Document not found: {doc_id}')
    
    # Generate new evaluation name from document(s)
    # Format: 
    #   1 doc:  "{Document Name} - MM/DD/YYYY"
    #   2 docs: "{Doc1} & {Doc2} - MM/DD/YYYY"
    #   3+ docs: "{First Doc} + {n} more - MM/DD/YYYY"
    
    new_eval_name = evaluation.get('name')  # Default to existing name
    doc_count = len(validated_doc_ids)
    
    if doc_count > 0:
        # Format date as MM/DD/YYYY (using local server time, not UTC)
        date_str = datetime.now().strftime('%m/%d/%Y')
        
        if doc_count == 1:
            # Single document
            first_doc = common.find_one('kb_docs', {'id': validated_doc_ids[0], 'kb_id': workspace_kb['id']})
            if first_doc:
                # Prefer filename over name (Supabase returns exact column names)
                doc_name = first_doc.get('filename') or first_doc.get('name', 'Document')
                doc_name_clean = remove_file_extension(doc_name)
                new_eval_name = f"{doc_name_clean} - {date_str}"
        
        elif doc_count == 2:
            # Two documents
            doc1 = common.find_one('kb_docs', {'id': validated_doc_ids[0], 'kb_id': workspace_kb['id']})
            doc2 = common.find_one('kb_docs', {'id': validated_doc_ids[1], 'kb_id': workspace_kb['id']})
            
            if doc1 and doc2:
                name1 = remove_file_extension(doc1.get('filename') or doc1.get('name', 'Doc1'))
                name2 = remove_file_extension(doc2.get('filename') or doc2.get('name', 'Doc2'))
                new_eval_name = f"{name1} & {name2} - {date_str}"
        
        else:
            # 3+ documents
            first_doc = common.find_one('kb_docs', {'id': validated_doc_ids[0], 'kb_id': workspace_kb['id']})
            if first_doc:
                doc_name = remove_file_extension(first_doc.get('filename') or first_doc.get('name', 'Document'))
                new_eval_name = f"{doc_name} + {doc_count - 1} more - {date_str}"
        
        logger.info(f"Auto-renaming eval {eval_id} ({doc_count} doc(s)): '{evaluation.get('name')}' -> '{new_eval_name}'")
    
    # Update evaluation record
    common.update_one(
        'eval_doc_summaries',
        {'id': eval_id},
        {
            'doc_type_id': doc_type_id,
            'criteria_set_id': criteria_set_id,
            'name': new_eval_name,
            'status': 'pending',
            'updated_by': user_id
        }
    )
    
    # Delete any existing doc set entries (in case of re-configuration)
    existing_doc_sets = common.find_many('eval_doc_sets', {'eval_summary_id': eval_id})
    for doc_set in existing_doc_sets:
        common.delete_one('eval_doc_sets', {'id': doc_set['id']})
    
    # Create new doc set entries
    for idx, doc_id in enumerate(validated_doc_ids):
        common.insert_one('eval_doc_sets', {
            'eval_summary_id': eval_id,
            'kb_doc_id': doc_id,
            'order_index': idx,
            'is_primary': idx == 0
        })
    
    # Send SQS message to trigger processing
    send_processing_message(
        eval_id=eval_id,
        org_id=org_id,
        workspace_id=workspace_id,
        doc_ids=validated_doc_ids,
        criteria_set_id=criteria_set_id
    )
    
    # Get updated evaluation for response
    updated_evaluation = common.find_one('eval_doc_summaries', {'id': eval_id})
    
    # Return updated evaluation
    result = common.format_record(updated_evaluation)
    result['docType'] = common.format_record(doc_type)
    result['criteriaSet'] = {
        'id': criteria_set['id'],
        'name': criteria_set['name'],
        'version': criteria_set.get('version')
    }
    result['docIds'] = validated_doc_ids
    
    return common.success_response(result)


def handle_delete_evaluation(eval_id: str, workspace_id: str, user_id: str) -> Dict[str, Any]:
    """Soft delete an evaluation."""
    eval_id = common.validate_uuid(eval_id, 'id')
    
    evaluation = common.find_one('eval_doc_summaries', {'id': eval_id, 'ws_id': workspace_id})
    if not evaluation:
        raise common.NotFoundError('Evaluation not found')
    
    # Soft delete
    common.update_one(
        'eval_doc_summaries',
        {'id': eval_id},
        {
            'is_deleted': True,
            'updated_by': user_id
        }
    )
    
    return common.success_response({
        'message': 'Evaluation deleted',
        'id': eval_id
    })


# =============================================================================
# RESULT EDITING HANDLERS
# =============================================================================

def handle_edit_result(
    event: Dict[str, Any],
    eval_id: str,
    result_id: str,
    workspace_id: str,
    user_id: str
) -> Dict[str, Any]:
    """
    Edit an evaluation result (narrative and/or status only).
    Creates a new edit record for version history.
    
    Request body:
    {
        "editedResult": "Updated explanation...",
        "editedStatusId": "uuid",
        "editNotes": "Reason for change..."
    }
    """
    eval_id = common.validate_uuid(eval_id, 'id')
    result_id = common.validate_uuid(result_id, 'resultId')
    
    # Verify evaluation exists and belongs to workspace
    evaluation = common.find_one('eval_doc_summaries', {'id': eval_id, 'ws_id': workspace_id})
    if not evaluation:
        raise common.NotFoundError('Evaluation not found')
    
    # Verify result exists and belongs to evaluation
    criteria_result = common.find_one('eval_criteria_results', {'id': result_id, 'eval_summary_id': eval_id})
    if not criteria_result:
        raise common.NotFoundError('Criteria result not found')
    
    body = json.loads(event.get('body', '{}'))
    
    edited_result = body.get('editedResult')
    edited_status_id = body.get('editedStatusId')
    edit_notes = body.get('editNotes')
    
    if not edited_result and not edited_status_id:
        raise common.ValidationError('Either editedResult or editedStatusId is required')
    
    # Get score_value from selected status option
    edited_score_value = None
    if edited_status_id:
        edited_status_id = common.validate_uuid(edited_status_id, 'editedStatusId')
        
        # Verify status option exists
        workspace = common.find_one('workspaces', {'id': workspace_id})
        org_id = workspace.get('org_id') if workspace else None
        
        status_options = get_status_options(org_id)
        status_option = next((s for s in status_options if s['id'] == edited_status_id), None)
        
        if not status_option:
            raise common.ValidationError('Invalid status option')
        
        # Capture score_value from status option
        if status_option.get('score_value') is not None:
            edited_score_value = float(status_option['score_value'])
    
    # Mark previous edits as not current
    previous_edits = common.find_many('eval_result_edits', {'criteria_result_id': result_id, 'is_current': True})
    for prev_edit in previous_edits:
        common.update_one('eval_result_edits', {'id': prev_edit['id']}, {'is_current': False})
    
    # Get the version number
    all_edits = common.find_many('eval_result_edits', {'criteria_result_id': result_id})
    version = len(all_edits) + 1
    
    # Create new edit record
    edit_data = {
        'criteria_result_id': result_id,
        'edited_result': edited_result,
        'edited_status_id': edited_status_id,
        'edited_score_value': edited_score_value,
        'edit_notes': edit_notes,
        'version': version,
        'is_current': True,
        'edited_by': user_id
    }
    
    edit = common.insert_one('eval_result_edits', edit_data)
    
    return common.success_response({
        'edit': common.format_record(edit),
        'criteriaResultId': result_id,
        'evaluationId': eval_id,
        'message': 'Result edited successfully'
    })


def handle_get_edit_history(
    eval_id: str,
    result_id: str,
    workspace_id: str,
    user_id: str
) -> Dict[str, Any]:
    """Get edit history for a criteria result."""
    eval_id = common.validate_uuid(eval_id, 'id')
    result_id = common.validate_uuid(result_id, 'resultId')
    
    # Verify evaluation exists and belongs to workspace
    evaluation = common.find_one('eval_doc_summaries', {'id': eval_id, 'ws_id': workspace_id})
    if not evaluation:
        raise common.NotFoundError('Evaluation not found')
    
    # Verify result exists
    criteria_result = common.find_one('eval_criteria_results', {'id': result_id, 'eval_summary_id': eval_id})
    if not criteria_result:
        raise common.NotFoundError('Criteria result not found')
    
    # Get all edits ordered by version
    edits = common.find_many(
        'eval_result_edits',
        {'criteria_result_id': result_id},
        order='version.desc'
    )
    
    # Format with editor info
    results = []
    for edit in edits:
        formatted = common.format_record(edit)
        
        # Get editor name
        if edit.get('edited_by'):
            editor = common.find_one('user_profiles', {'id': edit['edited_by']})
            if editor:
                formatted['editorName'] = f"{editor.get('first_name', '')} {editor.get('last_name', '')}".strip()
                formatted['editorEmail'] = editor.get('email')
        
        results.append(formatted)
    
    return common.success_response({
        'criteriaResultId': result_id,
        'evaluationId': eval_id,
        'history': results,
        'totalEdits': len(results)
    })


# =============================================================================
# EXPORT HANDLERS
# =============================================================================

def handle_export_pdf(
    eval_id: str,
    workspace_id: str,
    org_id: str,
    user_id: str
) -> Dict[str, Any]:
    """Export evaluation as PDF."""
    eval_id = common.validate_uuid(eval_id, 'id')
    
    # Get evaluation data
    evaluation = common.find_one('eval_doc_summaries', {'id': eval_id, 'ws_id': workspace_id})
    if not evaluation:
        raise common.NotFoundError('Evaluation not found')
    
    if evaluation.get('status') != 'completed':
        raise common.ValidationError('Cannot export incomplete evaluation')
    
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.units import inch
        
        # Create PDF in memory
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Title style
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30
        )
        
        # Add title
        story.append(Paragraph(f"Evaluation Report: {evaluation.get('name', 'Untitled')}", title_style))
        story.append(Spacer(1, 12))
        
        # Add summary info
        summary_data = [
            ['Status:', evaluation.get('status', 'Unknown').title()],
            ['Compliance Score:', f"{evaluation.get('compliance_score', 0):.1f}%"],
            ['Created:', str(evaluation.get('created_at', ''))[:19]],
            ['Completed:', str(evaluation.get('completed_at', ''))[:19] if evaluation.get('completed_at') else 'N/A']
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 4*inch])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 24))
        
        # Add document summary
        if evaluation.get('doc_summary'):
            story.append(Paragraph("Document Summary", styles['Heading2']))
            story.append(Paragraph(evaluation['doc_summary'], styles['Normal']))
            story.append(Spacer(1, 12))
        
        # Add evaluation summary
        if evaluation.get('eval_summary'):
            story.append(Paragraph("Evaluation Summary", styles['Heading2']))
            story.append(Paragraph(evaluation['eval_summary'], styles['Normal']))
            story.append(Spacer(1, 12))
        
        # Add criteria results
        story.append(Paragraph("Criteria Results", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        criteria_results = common.find_many(
            'eval_criteria_results',
            {'eval_summary_id': eval_id}
        )
        
        criteria_items = common.find_many(
            'eval_criteria_items',
            {'criteria_set_id': evaluation['criteria_set_id']},
            order='order_index.asc'
        )
        
        item_map = {item['id']: item for item in criteria_items}
        status_options = get_status_options(org_id)
        status_map = {s['id']: s for s in status_options}
        
        for result in criteria_results:
            item = item_map.get(result.get('criteria_item_id'), {})
            status = status_map.get(result.get('ai_status_id'), {})
            
            story.append(Paragraph(f"<b>{item.get('criteria_id', 'N/A')}: {item.get('requirement', 'Unknown')}</b>", styles['Normal']))
            story.append(Paragraph(f"Status: {status.get('name', 'Unknown')}", styles['Normal']))
            
            if result.get('ai_result'):
                story.append(Paragraph(f"Assessment: {result['ai_result'][:500]}", styles['Normal']))
            
            story.append(Spacer(1, 12))
        
        # Build PDF
        doc.build(story)
        
        # Get PDF content
        pdf_content = buffer.getvalue()
        buffer.close()
        
        # Return as base64 or upload to S3
        if EXPORT_BUCKET:
            # Upload to S3 and return presigned URL
            s3_key = f"exports/{workspace_id}/{eval_id}/report.pdf"
            url = upload_to_s3(pdf_content, s3_key, 'application/pdf')
            
            return common.success_response({
                'downloadUrl': url,
                'format': 'pdf',
                'evaluationId': eval_id
            })
        else:
            # Return base64 encoded content
            return common.success_response({
                'content': base64.b64encode(pdf_content).decode('utf-8'),
                'contentType': 'application/pdf',
                'fileName': f"evaluation-{eval_id[:8]}.pdf",
                'format': 'pdf'
            })
            
    except ImportError:
        logger.error("ReportLab not installed")
        raise common.ValidationError('PDF export not available (ReportLab not installed)')
    except Exception as e:
        logger.exception(f"Error generating PDF: {e}")
        raise common.ValidationError(f'Error generating PDF: {str(e)}')


def handle_export_xlsx(
    eval_id: str,
    workspace_id: str,
    org_id: str,
    user_id: str
) -> Dict[str, Any]:
    """Export evaluation as XLSX."""
    eval_id = common.validate_uuid(eval_id, 'id')
    
    # Get evaluation data
    evaluation = common.find_one('eval_doc_summaries', {'id': eval_id, 'ws_id': workspace_id})
    if not evaluation:
        raise common.NotFoundError('Evaluation not found')
    
    if evaluation.get('status') != 'completed':
        raise common.ValidationError('Cannot export incomplete evaluation')
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # ===============================
        # Sheet 1: Summary
        # ===============================
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Headers
        ws_summary['A1'] = "Evaluation Report"
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary.merge_cells('A1:D1')
        
        # Summary data
        summary_rows = [
            ('Name:', evaluation.get('name', 'Untitled')),
            ('Status:', evaluation.get('status', 'Unknown').title()),
            ('Compliance Score:', f"{evaluation.get('compliance_score', 0):.1f}%"),
            ('Created:', str(evaluation.get('created_at', ''))[:19]),
            ('Completed:', str(evaluation.get('completed_at', ''))[:19] if evaluation.get('completed_at') else 'N/A'),
            ('', ''),
            ('Document Summary:', ''),
            ('', evaluation.get('doc_summary', 'N/A')[:5000]),
            ('', ''),
            ('Evaluation Summary:', ''),
            ('', evaluation.get('eval_summary', 'N/A')[:5000])
        ]
        
        for row_idx, (label, value) in enumerate(summary_rows, start=3):
            ws_summary[f'A{row_idx}'] = label
            ws_summary[f'A{row_idx}'].font = Font(bold=True)
            ws_summary[f'B{row_idx}'] = value
        
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 80
        
        # ===============================
        # Sheet 2: Criteria Results
        # ===============================
        ws_results = wb.create_sheet("Criteria Results")
        
        # Headers
        headers = ['Criteria ID', 'Requirement', 'Category', 'Status', 'Confidence', 'Assessment', 'Weight']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_results.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Get results data
        criteria_results = common.find_many(
            'eval_criteria_results',
            {'eval_summary_id': eval_id}
        )
        
        criteria_items = common.find_many(
            'eval_criteria_items',
            {'criteria_set_id': evaluation['criteria_set_id']},
            order='order_index.asc'
        )
        
        item_map = {item['id']: item for item in criteria_items}
        status_options = get_status_options(org_id)
        status_map = {s['id']: s for s in status_options}
        
        for row_idx, result in enumerate(criteria_results, start=2):
            item = item_map.get(result.get('criteria_item_id'), {})
            status = status_map.get(result.get('ai_status_id'), {})
            
            ws_results.cell(row=row_idx, column=1, value=item.get('criteria_id', 'N/A'))
            ws_results.cell(row=row_idx, column=2, value=item.get('requirement', 'Unknown'))
            ws_results.cell(row=row_idx, column=3, value=item.get('category', ''))
            ws_results.cell(row=row_idx, column=4, value=status.get('name', 'Unknown'))
            ws_results.cell(row=row_idx, column=5, value=result.get('ai_confidence'))
            ws_results.cell(row=row_idx, column=6, value=result.get('ai_result', '')[:2000])
            ws_results.cell(row=row_idx, column=7, value=float(item.get('weight', 1.0)))
        
        # Adjust column widths
        ws_results.column_dimensions['A'].width = 15
        ws_results.column_dimensions['B'].width = 50
        ws_results.column_dimensions['C'].width = 20
        ws_results.column_dimensions['D'].width = 15
        ws_results.column_dimensions['E'].width = 12
        ws_results.column_dimensions['F'].width = 60
        ws_results.column_dimensions['G'].width = 10
        
        # ===============================
        # Sheet 3: Citations
        # ===============================
        ws_citations = wb.create_sheet("Citations")
        
        # Headers
        citation_headers = ['Criteria ID', 'Citation Text', 'Source']
        for col_idx, header in enumerate(citation_headers, start=1):
            cell = ws_citations.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        citation_row = 2
        for result in criteria_results:
            item = item_map.get(result.get('criteria_item_id'), {})
            citations = json.loads(result.get('ai_citations', '[]')) if result.get('ai_citations') else []
            
            for citation in citations:
                if isinstance(citation, str):
                    ws_citations.cell(row=citation_row, column=1, value=item.get('criteria_id', 'N/A'))
                    ws_citations.cell(row=citation_row, column=2, value=citation[:2000])
                    citation_row += 1
                elif isinstance(citation, dict):
                    ws_citations.cell(row=citation_row, column=1, value=item.get('criteria_id', 'N/A'))
                    ws_citations.cell(row=citation_row, column=2, value=citation.get('text', '')[:2000])
                    ws_citations.cell(row=citation_row, column=3, value=citation.get('source', ''))
                    citation_row += 1
        
        ws_citations.column_dimensions['A'].width = 15
        ws_citations.column_dimensions['B'].width = 80
        ws_citations.column_dimensions['C'].width = 30
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        xlsx_content = buffer.getvalue()
        buffer.close()
        
        # Return as base64 or upload to S3
        if EXPORT_BUCKET:
            s3_key = f"exports/{workspace_id}/{eval_id}/report.xlsx"
            url = upload_to_s3(xlsx_content, s3_key, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            
            return common.success_response({
                'downloadUrl': url,
                'format': 'xlsx',
                'evaluationId': eval_id
            })
        else:
            return common.success_response({
                'content': base64.b64encode(xlsx_content).decode('utf-8'),
                'contentType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'fileName': f"evaluation-{eval_id[:8]}.xlsx",
                'format': 'xlsx'
            })
            
    except ImportError:
        logger.error("openpyxl not installed")
        raise common.ValidationError('XLSX export not available (openpyxl not installed)')
    except Exception as e:
        logger.exception(f"Error generating XLSX: {e}")
        raise common.ValidationError(f'Error generating XLSX: {str(e)}')


# =============================================================================
# SQS INTEGRATION
# =============================================================================

def send_processing_message(
    eval_id: str,
    org_id: str,
    workspace_id: str,
    doc_ids: List[str],
    criteria_set_id: str
) -> None:
    """Send message to SQS to trigger async processing."""
    if not SQS_QUEUE_URL:
        logger.warning("SQS_QUEUE_URL not configured, skipping message send")
        return
    
    try:
        sqs = boto3.client('sqs', region_name=AWS_REGION)
        
        message = {
            'eval_id': eval_id,
            'org_id': org_id,
            'ws_id': workspace_id,
            'doc_ids': doc_ids,
            'criteria_set_id': criteria_set_id,
            'action': 'evaluate'
        }
        
        # Build SQS params conditionally (boto3 rejects None for optional params)
        sqs_params = {
            'QueueUrl': SQS_QUEUE_URL,
            'MessageBody': json.dumps(message)
        }
        
        # Only add MessageGroupId for FIFO queues
        if '.fifo' in SQS_QUEUE_URL:
            sqs_params['MessageGroupId'] = workspace_id
        
        sqs.send_message(**sqs_params)
        
        logger.info(f"Sent processing message for eval {eval_id}")
        
    except Exception as e:
        logger.error(f"Error sending SQS message: {e}")
        # Don't fail the request, just log the error
        # The evaluation is created, user can retry processing


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def remove_file_extension(filename: str) -> str:
    """
    Remove common file extensions from filename.
    
    Strips extensions like .pdf, .doc, .docx, .txt, .xlsx, etc.
    Returns the filename without extension.
    """
    if not filename:
        return filename
    
    # Common file extensions to remove
    extensions = [
        '.pdf', '.doc', '.docx', '.txt', '.rtf',
        '.xlsx', '.xls', '.csv',
        '.ppt', '.pptx',
        '.zip', '.tar', '.gz',
        '.jpg', '.jpeg', '.png', '.gif',
        '.html', '.htm', '.xml', '.json'
    ]
    
    filename_lower = filename.lower()
    for ext in extensions:
        if filename_lower.endswith(ext):
            return filename[:-len(ext)]
    
    return filename


def get_effective_eval_config(org_id: Optional[str]) -> Dict[str, Any]:
    """
    Get effective evaluation configuration for an organization.
    Org config overrides system config when present.
    
    Returns:
        {
            'categorical_mode': 'boolean' | 'detailed',
            'show_numerical_score': bool
        }
    """
    # Get system config (should always exist with defaults)
    sys_config = common.find_one('eval_cfg_sys', {})
    if not sys_config:
        # Fallback defaults if no system config exists
        logger.warning("No system eval config found, using defaults")
        return {
            'categorical_mode': 'detailed',
            'show_numerical_score': True
        }
    
    # If no org_id, return system config
    if not org_id:
        return {
            'categorical_mode': sys_config.get('categorical_mode', 'detailed'),
            'show_numerical_score': sys_config.get('show_numerical_score', True)
        }
    
    # Get org config (may not exist)
    org_config = common.find_one('eval_cfg_org', {'org_id': org_id})
    
    # Apply precedence: org overrides sys when not null
    effective_config = {
        'categorical_mode': (
            org_config.get('categorical_mode') if org_config and org_config.get('categorical_mode') is not None
            else sys_config.get('categorical_mode', 'detailed')
        ),
        'show_numerical_score': (
            org_config.get('show_numerical_score') if org_config and org_config.get('show_numerical_score') is not None
            else sys_config.get('show_numerical_score', True)
        )
    }
    
    return effective_config


def get_status_options(org_id: Optional[str], categorical_mode: Optional[str] = None) -> List[Dict[str, Any]]:
    """
    Get active status options for an organization, filtered by categorical_mode.
    
    Args:
        org_id: Organization ID (None for system defaults)
        categorical_mode: Filter by mode ('boolean' or 'detailed'). If None, returns all modes.
    
    Returns:
        List of status options matching the mode
    """
    if org_id:
        # Check for org-level options
        filters = {'org_id': org_id}
        
        org_options = common.find_many(
            'eval_org_status_options',
            filters,
            order='order_index.asc'
        )
        if org_options:
            # Filter by mode if specified
            if categorical_mode:
                org_options = [
                    opt for opt in org_options 
                    if opt.get('mode') == categorical_mode or opt.get('mode') == 'both'
                ]
            return org_options
    
    # Fall back to system options
    filters = {}
    sys_options = common.find_many(
        'eval_sys_status_options',
        filters,
        order='order_index.asc'
    )
    
    # Filter by mode if specified
    if categorical_mode:
        sys_options = [
            opt for opt in sys_options 
            if opt.get('mode') == categorical_mode or opt.get('mode') == 'both'
        ]
    
    return sys_options


def upload_to_s3(content: bytes, key: str, content_type: str) -> str:
    """Upload content to S3 and return presigned URL."""
    try:
        s3 = boto3.client('s3', region_name=AWS_REGION)
        
        s3.put_object(
            Bucket=EXPORT_BUCKET,
            Key=key,
            Body=content,
            ContentType=content_type
        )
        
        # Generate presigned URL (valid for 1 hour)
        url = s3.generate_presigned_url(
            'get_object',
            Params={'Bucket': EXPORT_BUCKET, 'Key': key},
            ExpiresIn=3600
        )
        
        return url
        
    except Exception as e:
        logger.error(f"Error uploading to S3: {e}")
        raise
